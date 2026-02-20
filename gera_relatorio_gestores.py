
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Define paths
INPUT_FILE = r"d:\APRENDIZADO APP\MEDICOES\CONTROLES POR COMISSÃO E GESTORES.xlsx"
OUTPUT_FILE = r"d:\APRENDIZADO APP\MEDICOES\RELATORIO DE OBRAS POR GESTORES E FISCAIS.xlsx"

def load_data(file_path):
    print(f"Reading {file_path}...")
    # Load the entire sheet to parse manually due to multi-header structure
    try:
        df_raw = pd.read_excel(file_path, header=None)
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return None

    data_rows = []
    current_region = None
    
    # Iterate rows to find blocks
    # Structure:
    # Row N: Region Name (BAIXADA, etc)
    # Row N+1: Headers (SEI, GESTOR, etc)
    # Row N+M: Data...
    
    # We look for the "SEI" header to identify start of data
    header_row_idx = -1
    cols_map = {}
    
    # Standard headers we expect
    expected_headers = ['SEI', 'GESTOR(A) ATUANTE', 'FISCAL NOMEADO', 'MUNICIPIO', 'EMPRESA', '%EXEC', 'STATUS']
    
    i = 0
    while i < len(df_raw):
        row = df_raw.iloc[i].astype(str).tolist()
        # Clean row values: handle newlines and strip
        row_clean = [str(v).replace('\n', ' ').strip().upper() for v in row]
        
        # Check if this is a header row
        # Relaxed check: Look for "SEI" and "GESTOR" part
        has_sei = "SEI" in row_clean
        has_gestor = any("GESTOR" in c for c in row_clean)
        
        if has_sei and has_gestor:
            # Found a header row
            print(f"DEBUG: Found header at row {i}")
            # The region should be in the row above (i-1) in column 0 or 1
            if i > 0:
                possible_region = str(df_raw.iloc[i-1, 0]).strip()
                if possible_region.lower() == 'nan' or not possible_region:
                     possible_region = str(df_raw.iloc[i-1, 1]).strip()
                
                # If valid region text, use it. Otherwise keep previous or default.
                if possible_region.upper() not in ['NAN', '', 'TOTAL']:
                     current_region = possible_region.upper()
            
            # Map columns
            cols_map = {}
            for idx, val in enumerate(row_clean):
                 if val:
                    cols_map[val] = idx
            
            print(f"DEBUG: Columns found: {list(cols_map.keys())}")

            # Process data lines until empty line or "Total"
            j = i + 1
            while j < len(df_raw):
                d_row = df_raw.iloc[j]
                d_row_str = [str(v).strip() for v in d_row]
                
                # Check for end of block
                first_val = d_row_str[0].upper() if len(d_row_str)>0 else ""
                second_val = d_row_str[1].upper() if len(d_row_str)>1 else ""
                
                if not first_val and 'TOTAL' in second_val:
                    break # End of block
                if first_val in ['BAIXADA', 'SUL', 'NORTE', 'METROPOLITANA', 'CENTRO'] and j > i+1: 
                    # If we encounter a region name in col 0, it's likely a header for next block
                    break 
                
                # Check if it has data (SEI usually)
                if 'SEI' in cols_map:
                    sei_idx = cols_map['SEI']
                    sei_val = str(d_row[sei_idx]).strip()
                    
                    if sei_val and sei_val.lower() != 'nan' and sei_val.upper() != 'SEI' and "TOTAL" not in sei_val.upper():
                        # Extract values
                        record = {'REGIÃO': current_region}
                        for head, col_idx in cols_map.items():
                            val = d_row[col_idx]
                            record[head] = val
                        data_rows.append(record)
                
                j += 1
            i = j # Move outer loop
        else:
            i += 1
            
    return pd.DataFrame(data_rows)

def generate_report(df):
    if df is None or df.empty:
        print("No data found to generate report.")
        return

    # Clean Data
    df['GESTOR(A) ATUANTE'] = df['GESTOR(A) ATUANTE'].fillna('NÃO DEFINIDO').astype(str).str.strip().str.upper()
    df['FISCAL NOMEADO'] = df['FISCAL NOMEADO'].fillna('NÃO DEFINIDO').astype(str).str.strip().str.upper()
    df['REGIÃO'] = df['REGIÃO'].fillna('SEM REGIÃO').astype(str).str.strip().str.upper()

    # Separate df into expanded sets taking into account "/" separated multiple people
    df_f = df.assign(FISCAL_INDIVIDUAL=df['FISCAL NOMEADO'].str.split('/')).explode('FISCAL_INDIVIDUAL')
    df_f['FISCAL_INDIVIDUAL'] = df_f['FISCAL_INDIVIDUAL'].str.strip()
    df_f = df_f[~df_f['FISCAL_INDIVIDUAL'].isin(['', 'NAN'])]
    
    df_g = df.assign(GESTOR_INDIVIDUAL=df['GESTOR(A) ATUANTE'].str.split('/')).explode('GESTOR_INDIVIDUAL')
    df_g['GESTOR_INDIVIDUAL'] = df_g['GESTOR_INDIVIDUAL'].str.strip()
    df_g = df_g[~df_g['GESTOR_INDIVIDUAL'].isin(['', 'NAN'])]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo Obras"

    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells("A1:G2")
    cell = ws.cell(row=1, column=1, value="RELATÓRIO DE OBRAS POR GESTORES E FISCAIS")
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Title border
    for r in range(1, 3):
        for c in range(1, 8):
            if c != 4:  # col D is spacing
                ws.cell(row=r, column=c).border = thin_border
            
    current_row = 4
    
    regions = df['REGIÃO'].unique()
    regions = sorted([r for r in regions if r.upper() not in ['NAN', '', 'SEM REGIÃO']])
    if 'SEM REGIÃO' in df['REGIÃO'].unique():
        regions.append('SEM REGIÃO')
        
    for region in regions:
        # Get data for region
        rf = df_f[df_f['REGIÃO'] == region]
        rg = df_g[df_g['REGIÃO'] == region]
        
        # Aggregate Fiscais
        fiscal_counts = rf['FISCAL_INDIVIDUAL'].value_counts().reset_index()
        fiscal_counts.columns = ['Nome', 'Count']
        fiscal_counts = fiscal_counts.sort_values(by='Nome')
        
        # Aggregate Gestores
        gestor_counts = rg['GESTOR_INDIVIDUAL'].value_counts().reset_index()
        gestor_counts.columns = ['Nome', 'Count']
        gestor_counts = gestor_counts.sort_values(by='Nome')
        
        # Skip if no data
        if fiscal_counts.empty and gestor_counts.empty:
            continue
            
        # Region Header for Fiscais
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        h_f = ws.cell(row=current_row, column=1, value=f"{region} - Obras por FISCAL")
        h_f.font = header_font
        h_f.alignment = Alignment(horizontal='center')
        
        for c in range(1, 4):
            ws.cell(row=current_row, column=c).border = thin_border
            
        # Region Header for Gestores
        ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=7)
        h_g = ws.cell(row=current_row, column=5, value=f"{region} - Obras por GESTOR")
        h_g.font = header_font
        h_g.alignment = Alignment(horizontal='center')
        for c in range(5, 8):
            ws.cell(row=current_row, column=c).border = thin_border
        
        current_row += 1
        
        max_rows = max(len(fiscal_counts), len(gestor_counts))
        
        for i in range(max_rows):
            # Fiscais Columns
            if i < len(fiscal_counts):
                f_row = fiscal_counts.iloc[i]
                ws.cell(row=current_row + i, column=1, value=i+1).border = thin_border
                ws.cell(row=current_row + i, column=2, value=f_row['Nome']).border = thin_border
                ws.cell(row=current_row + i, column=3, value=int(f_row['Count'])).border = thin_border
                
                # Alignments
                ws.cell(row=current_row + i, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=current_row + i, column=3).alignment = Alignment(horizontal='center')
            else:
                # Fill empty cells with border to keep table rectangular
                for c in range(1, 4):
                    ws.cell(row=current_row + i, column=c).border = thin_border
            
            # Gestores Columns
            if i < len(gestor_counts):
                g_row = gestor_counts.iloc[i]
                ws.cell(row=current_row + i, column=5, value=i+1).border = thin_border
                ws.cell(row=current_row + i, column=6, value=g_row['Nome']).border = thin_border
                ws.cell(row=current_row + i, column=7, value=int(g_row['Count'])).border = thin_border
                
                ws.cell(row=current_row + i, column=5).alignment = Alignment(horizontal='center')
                ws.cell(row=current_row + i, column=7).alignment = Alignment(horizontal='center')
            else:
                # Fill empty cells with border
                for c in range(5, 8):
                    ws.cell(row=current_row + i, column=c).border = thin_border
                
        current_row += max_rows
        
        current_row += 3  # Add spacing before next region
        
    # --- QUADRO RESUMO GERAL ---
    current_row += 1
    
    # Calcular contagem global (ignorando região)
    overall_fiscal = df_f['FISCAL_INDIVIDUAL'].value_counts().reset_index()
    overall_fiscal.columns = ['Nome', 'Count']
    overall_fiscal = overall_fiscal.sort_values(by='Nome')
    
    overall_gestor = df_g['GESTOR_INDIVIDUAL'].value_counts().reset_index()
    overall_gestor.columns = ['Nome', 'Count']
    overall_gestor = overall_gestor.sort_values(by='Nome')
    
    # Headers do Quadro Geral
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    h_f = ws.cell(row=current_row, column=1, value="FISCAL")
    h_f.font = header_font
    h_f.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=3, value="OBRAS").font = header_font
    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=current_row, start_column=5, end_row=current_row, end_column=6)
    h_g = ws.cell(row=current_row, column=5, value="GESTOR")
    h_g.font = header_font
    h_g.alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=7, value="OBRAS").font = header_font
    ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='center')
    
    for c in list(range(1, 4)) + list(range(5, 8)):
        ws.cell(row=current_row, column=c).border = thin_border
        
    current_row += 1
    
    max_overall = max(len(overall_fiscal), len(overall_gestor))
    
    for i in range(max_overall):
        # Lado Esquerdo (Fiscal)
        ws.merge_cells(start_row=current_row + i, start_column=1, end_row=current_row + i, end_column=2)
        if i < len(overall_fiscal):
            f_row = overall_fiscal.iloc[i]
            c_f = ws.cell(row=current_row + i, column=1, value=f_row['Nome'])
            c_f.alignment = Alignment(horizontal='left')
            ws.cell(row=current_row + i, column=3, value=int(f_row['Count'])).alignment = Alignment(horizontal='center')
            
        for c in range(1, 4):
            ws.cell(row=current_row + i, column=c).border = thin_border
            
        # Lado Direito (Gestor)
        ws.merge_cells(start_row=current_row + i, start_column=5, end_row=current_row + i, end_column=6)
        if i < len(overall_gestor):
            g_row = overall_gestor.iloc[i]
            c_g = ws.cell(row=current_row + i, column=5, value=g_row['Nome'])
            c_g.alignment = Alignment(horizontal='left')
            ws.cell(row=current_row + i, column=7, value=int(g_row['Count'])).alignment = Alignment(horizontal='center')
            
        for c in range(5, 8):
            ws.cell(row=current_row + i, column=c).border = thin_border

    current_row += max_overall
    
    current_row += 3
    # --- FIM QUADRO RESUMO GERAL ---

    # Set fixed column widths for the layout
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 5   # Spacing column
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 12
        
    wb.save(OUTPUT_FILE)
    print(f"Report generated: {OUTPUT_FILE}")

if __name__ == "__main__":
    df = load_data(INPUT_FILE)
    if df is not None:
        generate_report(df)
