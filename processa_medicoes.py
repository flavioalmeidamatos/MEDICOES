from typing import Any # type: ignore
import pandas as pd # type: ignore
import openpyxl # type: ignore
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side # type: ignore
from openpyxl.utils import get_column_letter # type: ignore
from datetime import datetime
import os
import re

# Caminhos dos arquivos
# Usa o diretório do próprio script para funcionar tanto no Windows quanto aqui.
CWD = os.path.dirname(os.path.abspath(__file__))
FILE_BASE = os.path.join(CWD, "BASE.xlsx")
FILE_ANALITICA = os.path.join(CWD, "ANALITICA.xlsx")
FILE_AUXILIAR = os.path.join(CWD, "AUXILIAR.xlsx")
FILE_COMISSOES = os.path.join(CWD, "COMISSÕES POR REGIAO.xlsx")
FILE_CONTROLES = os.path.join(CWD, "CONTROLES POR COMISSÃO E GESTORES.xlsx")
FILE_OUTPUT = os.path.join(CWD, "MEDIÇÕES_CONSOLIDADO.xlsx")

def clean_sei(val):
    if pd.isna(val): return ""
    return str(val).strip()

def to_numeric(val):
    if pd.isna(val): return 0.0
    if isinstance(val, (int, float)): return float(val)
    # Remove R$, espaços, pontos de milhar, troca vírgula por ponto
    s = str(val).replace("R$", "").replace("\xa0", "").replace(" ", "")
    # Se houver pontos e vírgulas, assume que ponto é milhar e vírgula é decimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        f_val: float = float(s)
        return float(round(f_val, 2)) # type: ignore
    except:
        return 0.0

def get_region_mapping():
    # Lê AUXILIAR.xlsx para mapear Município -> Região
    df_aux = pd.read_excel(FILE_AUXILIAR, sheet_name="AUXILIAR")
    mapping = {}
    siglas = {
        "BAIXADA": "BX",
        "METROPOLITANA": "MT",
        "SUL FLUMINENSE": "SL",
        "NORTE": "NT"
    }
    # O arquivo AUXILIAR tem colunas com nomes das Regiões e os municípios abaixo
    for col in df_aux.columns:
        if "Unnamed" in col: continue
        reg_name = col.strip().upper()
        if reg_name in siglas:
            sigla = siglas[reg_name]
            for muni in df_aux[col].dropna():
                mapping[str(muni).strip().upper()] = sigla
    return mapping

def normalize_name(name):
    if not name or pd.isna(name): return ""
    # Remove pontos, traços, barras e espaços múltiplos para comparação
    n = str(name).upper().strip()
    n = re.sub(r'[\.\-\/]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n

def get_contractor_mapping():
    # Lê AUXILIAR.xlsx para mapear CONTRATADA -> RESUMIDO
    df_aux = pd.read_excel(FILE_AUXILIAR, sheet_name="AUXILIAR")
    mapping = {}
    if 'CONTRATADA' in df_aux.columns and 'RESUMIDO' in df_aux.columns:
        for _, row in df_aux[['CONTRATADA', 'RESUMIDO']].dropna(subset=['CONTRATADA', 'RESUMIDO']).iterrows():
            orig = normalize_name(row['CONTRATADA'])
            res = str(row['RESUMIDO']).strip()
            if orig:
                mapping[orig] = res
    return mapping

def get_concluidas_sei() -> Any:
    # Lê AUXILIAR.xlsx para obter lista de SEIs que devem ser tratados como CONCLUÍDOS
    # (Tabela "CONCLUIDAS" mencionada - coluna SEI no arquivo AUXILIAR)
    try:
        df_aux = pd.read_excel(FILE_AUXILIAR, sheet_name="AUXILIAR")
        if 'SEI' in df_aux.columns:
            # Pega todos os SEIs da coluna, limpa e retorna como um set
            concluidas = df_aux['SEI'].dropna().apply(clean_sei).unique()
            return set(concluidas)
    except Exception as e:
        print(f"Erro ao ler SEIs concluídos de AUXILIAR.xlsx: {e}")
    return set()

def get_comissoes_data():
    xl = pd.ExcelFile(FILE_COMISSOES)
    data = {}

    # --- PASSO 1: Lê a aba AUXILIAR para STATUS e LOCAL (gestor aqui é apenas fallback) ---
    aux_sheet = next((s for s in xl.sheet_names if s.upper() == "AUXILIAR"), None)
    if aux_sheet:
        df_aux = pd.read_excel(FILE_COMISSOES, sheet_name=aux_sheet)
        df_aux.columns = [str(c).replace("\n", " ").upper().strip() for c in df_aux.columns]
        if 'SEI' in df_aux.columns:
            for _, row in df_aux.iterrows():
                sei = clean_sei(row['SEI'])
                if not sei: continue
                # Tenta ler gestor da AUXILIAR como fallback inicial
                gestor_aux = ""
                for col_g in ['GESTOR(A) ATUANTE', 'GESTOR(A)', 'GESTOR']:
                    if col_g in df_aux.columns and pd.notna(row.get(col_g)): # type: ignore
                        gestor_aux = str(row[col_g]).strip() # type: ignore
                        break
                status_val = str(row['STATUS']).replace('#', '').strip().upper() if 'STATUS' in df_aux.columns and pd.notna(row.get('STATUS')) else "" # type: ignore
                local_val  = str(row['LOCAL']).strip().upper() if 'LOCAL' in df_aux.columns and pd.notna(row.get('LOCAL')) else "CIVIS" # type: ignore
                data[sei] = {
                    'gestor': gestor_aux,
                    'status_aux': status_val,
                    'local': local_val
                }

    # --- PASSO 2: Abas regionais — fonte primária do GESTOR(A) ATUANTE por SEI ---
    for sheet in xl.sheet_names:
        if sheet.upper() == "AUXILIAR":
            continue

        local_val = "CIVIS"
        if sheet.upper() == "CONTIGENCIA":
            local_val = "CONTINGENCIA"
        elif sheet.upper() == "ESPECIAIS":
            local_val = "ESPECIAIS"

        df = pd.read_excel(FILE_COMISSOES, sheet_name=sheet, header=None)

        sei_idx    = None
        gestor_idx = None
        start_row  = 0

        # Detecta linha de cabeçalho e índices das colunas SEI e GESTOR(A) ATUANTE
        for i in range(min(10, len(df))):
            row_vals = df.iloc[i].fillna("").astype(str).str.upper().tolist()
            s, g = None, None
            for j, val in enumerate(row_vals):
                v = val.replace("\n", " ").strip()
                if v in ("SEI", "PROCESSO SEI"):
                    s = j
                # Aceita qualquer variação: GESTOR(A) ATUANTE, GESTOR ATUANTE, GESTOR(A), GESTOR
                if any(x in v for x in ["GESTOR(A) ATUANTE", "GESTOR ATUANTE", "GESTOR(A)", "GESTOR"]):
                    g = j
            if s is not None and g is not None:
                sei_idx, gestor_idx, start_row = s, g, i + 1
                break

        if sei_idx is None:
            print(f"  Aviso: coluna SEI não encontrada na aba '{sheet}' — pulando.")
            continue

        for i in range(start_row, len(df)):
            row = df.iloc[i]
            sei = clean_sei(row[sei_idx])
            if not sei or sei.upper() == "NAN":
                continue

            # Lê o GESTOR(A) ATUANTE da aba regional
            gestor_regional = ""
            if gestor_idx is not None and pd.notna(row[gestor_idx]):
                gestor_regional = str(row[gestor_idx]).strip()

            if sei not in data:
                # Novo registro: cria com dados da aba regional
                data[sei] = {
                    'gestor': gestor_regional,
                    'local': local_val,
                    'status_aux': ''
                }
            else:
                # Já existia (da AUXILIAR): atualiza LOCAL e SOBRESCREVE gestor
                # com o GESTOR(A) ATUANTE da aba regional — fonte mais confiável
                data[sei]['local'] = local_val # type: ignore
                if gestor_regional:
                    data[sei]['gestor'] = gestor_regional # type: ignore

    return data

def get_gestor_fiscal_data():
    """Unifica dados de GESTOR e FISCAL dos dois arquivos de controles/comissões."""
    # 1. Dados do COMISSÕES POR REGIAO.xlsx (Fonte tradicional)
    data = get_comissoes_data()
    # Adiciona fiscal inicial vazio
    for sei in data:
        data[sei]['fiscal'] = ""

    # 2. Dados do CONTROLES POR COMISSÃO E GESTORES.xlsx (Fonte mais atualizada/detalhada)
    if os.path.exists(FILE_CONTROLES):
        try:
            # Lê todas as tabelas ou o sheet Planilha1
            df_ctrl_raw = pd.read_excel(FILE_CONTROLES, header=None)
            
            # Percorre o arquivo buscando blocos de dados (SEI e GESTOR)
            for i in range(len(df_ctrl_raw)): # type: ignore
                # Limpa novos nomes de colunas (remove \n e espaços extras)
                row_vals = [str(v).replace("\n", " ").strip().upper() for v in df_ctrl_raw.iloc[i].fillna("").astype(str)] # type: ignore
                if "SEI" in row_vals and any("GESTOR" in v for v in row_vals):
                    # Achou cabeçalho
                    cols = {v: idx for idx, v in enumerate(row_vals) if v}
                    
                    # Identifica índices exatos para evitar fallback para 0 (SEI)
                    sei_idx = cols.get("SEI")
                    gestor_idx = None
                    for gk in ["GESTOR(A) ATUANTE", "GESTOR ATUANTE", "GESTOR(A)", "GESTOR"]:
                        if gk in cols:
                            gestor_idx = cols[gk]
                            break
                    
                    if sei_idx is None or gestor_idx is None:
                        continue

                    # Processa linhas abaixo até encontrar vazio
                    for j in range(i + 1, len(df_ctrl_raw)):
                        d_row = df_ctrl_raw.iloc[j]
                        sei_orig = str(d_row[sei_idx]).strip()
                        sei = clean_sei(sei_orig)
                        if not sei or sei.upper() == "NAN" or "TOTAL" in sei.upper():
                            if not sei_orig or sei_orig.upper() == "NAN": break
                            else: continue
                        
                        gestor = str(d_row[gestor_idx]).strip()
                        fiscal = ""
                        for fk in ["FISCAL NOMEADO", "FISCAL", "FISCAL(A)"]:
                            if fk in cols:
                                fiscal = str(d_row[cols[fk]]).strip()
                                break
                        
                        status_val = ""
                        if "STATUS" in cols:
                            status_val = str(d_row[cols["STATUS"]]).strip().upper().replace("#", "")

                        if sei not in data:
                            data[sei] = {'gestor': gestor, 'fiscal': fiscal, 'local': 'CIVIS', 'status_aux': status_val} # type: ignore
                        else:
                            # Prioriza dados do arquivo de CONTROLES se preenchidos
                            if gestor and gestor.upper() != "NAN": data[sei]['gestor'] = gestor # type: ignore
                            if fiscal and fiscal.upper() != "NAN": data[sei]['fiscal'] = fiscal # type: ignore
                            if status_val: data[sei]['status_aux'] = status_val # type: ignore
        except Exception as e:
            print(f"Erro ao ler arquivo de controles: {e}")

    return data

def apply_sheet_formatting(ws, col_map, header, all_months, model_widths, model_header_style,
                           h_vlr_contr, h_med_acum, h_saldo, h_inicio):
    """Aplica formatação idêntica (cabeçalhos, cores, bordas, larguras) a uma worksheet."""
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Fill colors
    fill_header = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    # Colors for LOCAL (vibrantes conforme imagem)
    fills_local = {
        "CIVIS": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),        # Laranja
        "CONTINGENCIA": PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),  # Amarelo
        "ESPECIAIS": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")     # Verde Água
    }

    fills_regiao = {
        "SL": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "NT": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "BX": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "MT": PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    }

    # Header format
    for cell in ws[1]:
        name_in_cell = str(cell.value).replace('\n', ' ').strip()

        # Estilo base
        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        cell.border = thin_border

        # Tenta aplicar do modelo
        if name_in_cell in model_header_style:
            style = model_header_style[name_in_cell]
            if style['fill'] and style['fill'] != '00000000':
                cell.fill = PatternFill(start_color=style['fill'], end_color=style['fill'], fill_type="solid")
            else:
                cell.fill = fill_header  # Fallback cinza
            cell.font = Font(bold=style['font_bold'], color=style['font_color'])
        else:
            cell.fill = fill_header
            cell.font = Font(bold=True)

    # Data content
    money_fmt = '_-R$ * #,##0.00_-;_-R$ * -#,##0.00_-;_-R$ * "-"??_-;_-@_-'
    ws_any: Any = ws
    
    for row in range(2, ws_any.max_row + 1):
        for col in range(1, len(header) + 1):
            ws_any.cell(row=row, column=col).border = thin_border

        # Formatação LOCAL
        if "LOCAL" in col_map:
            local_cell = ws_any.cell(row=row, column=col_map["LOCAL"])
            local_val = str(local_cell.value).strip().upper()
            if local_val in fills_local:
                local_cell.fill = fills_local[local_val]
                local_cell.font = Font(bold=True)

        # Formatação REGIÃO
        if "REGIÃO" in col_map:
            reg_val = ws_any.cell(row=row, column=col_map["REGIÃO"]).value
            if reg_val in fills_regiao:
                ws_any.cell(row=row, column=col_map["REGIÃO"]).fill = fills_regiao[reg_val]

        # Formatação Financeira (heurística + lista passada)
        # Se o nome da coluna estiver em all_months ou contiver VLR, SALDO, MEDIÇÕES (exceto ano solto)
        for col_name, col_idx in col_map.items():
            is_money = False
            # Verifica meses
            c_clean = col_name.replace(" ", "")
            if re.match(r'^[A-Z]{3}/\d{2}$', c_clean):
                is_money = True
            elif any(k in col_name.upper() for k in ["VLR", "VALOR", "SALDO", "MEDIÇÕES", "MEDICOES"]):
                # Evita "MEDIÇÕES 2025" se for só contagem, mas aqui é valor, então ok
                is_money = True
            
            if is_money:
                cell_val: Any = ws_any.cell(row=row, column=col_idx) # type: ignore
                cell_val.number_format = money_fmt # type: ignore
                try:
                    if cell_val.value is not None: # type: ignore
                        val_clean = str(cell_val.value).replace('R$', '').replace(' ', '') # type: ignore
                        if ',' in val_clean and '.' not in val_clean:
                            val_clean = val_clean.replace(',', '.')
                        cell_val.value = float(round(float(val_clean), 2)) # type: ignore
                    else:
                        cell_val.value = 0.0
                except:
                    cell_val.value = 0.0

        # Formatação de Datas
        for dc in [h_inicio, "DATA FINAL", "Prazo Final", "Ordem de Início"]:
            if dc in col_map:
                cell_dt: Any = ws_any.cell(row=row, column=col_map[dc]) # type: ignore
                if cell_dt.value: # type: ignore
                    cell_dt.number_format = 'DD/MM/YYYY' # type: ignore
        
        if "% EXEC." in col_map:
            ws_any.cell(row=row, column=col_map["% EXEC."]).number_format = '0.00%' # type: ignore

    # --- Aplicação de Larguras ---
    for col in ws.columns:
        column_letter = col[0].column_letter
        val_header = str(col[0].value).replace('\n', ' ').strip()
        
        # Base width from model or fallback
        width = model_widths.get(val_header, 15)
        
        # Sanity checks and adjustments
        # 1. Currency Columns (Months and Totals)
        is_money = False
        if re.match(r'^[A-Z]{3}/\d{2}$', val_header):
            is_money = True
        elif any(k in val_header.upper() for k in ["VLR", "VALOR", "SALDO", "MEDIÇÕES", "MEDICOES"]):
            is_money = True
        
        if is_money:
            # Mantemos 20 para meses normais, mas permitimos exceções abaixo
            width = max(width, 20)
        
        # 2. Date Columns
        if any(k in val_header.upper() for k in ["DATA", "INÍCIO", "FINAL"]):
            width = max(width, 12)
            
        # 3. Specific Columns (Refinamento conforme pedido)
        if val_header == "Nº":
            width = 4
        elif val_header == "SEI":
            width = 19
        elif "PRAZO" in val_header.upper():
            width = 11
        elif val_header == "VLR.CONTRATO C/ADITIVO":
            width = 18
        elif val_header == "GESTOR":
            width = 19
        elif val_header == "FISCAL":
            width = 25
        elif val_header == "MEDIÇÕES 2025":
            width = 18
        elif val_header == "MEDIÇÕES 2026":
            width = 18
        elif val_header == "MEDIÇÕES ACUMULADAS":
            width = 18
        elif val_header == "SALDO DO CONTRATO":
            width = 18
        elif val_header == "MUNICIPIO":
            width = min(max(width, 15), 22)
        elif val_header == "CONTRATADA":
            width = max(width, 18)
        elif "%" in val_header:
            width = 8
            
        ws.column_dimensions[column_letter].width = width


def prepare_dataframe(df, keep_execution=True):
    """Filtra, ordena e numera o DataFrame conforme os status desejados."""
    # Filtro flexível para incluir variações como "ATA DE REGISTRO DE PREÇO"
    status_col = df['STATUS'].astype(str).str.upper()
    mask = status_col.str.contains('EXECUÇÃO|EXECUCAO|ATA DE REGISTRO', na=False)
    
    if keep_execution:
        df_filtered = df[mask].copy()
    else:
        df_filtered = df[~mask].copy()

    # Remover duplicatas residuais
    df_filtered = df_filtered.drop_duplicates(subset=['SEI']).copy()

    # Definir ordem customizada para LOCAL
    local_rank = {"CIVIS": 0, "CONTINGENCIA": 1, "ESPECIAIS": 2}
    df_filtered['_rank_local'] = df_filtered['LOCAL'].map(lambda x: local_rank.get(str(x).upper().strip(), 99))

    # Ordenar
    df_filtered = df_filtered.sort_values(by=["_rank_local", "CONTRATADA"], ascending=[True, True]).reset_index(drop=True)
    df_filtered = df_filtered.drop(columns=['_rank_local'])

    # Numerar sequencialmente
    if "Nº" in df_filtered.columns:
        df_filtered = df_filtered.drop(columns=["Nº"])
    df_filtered.insert(0, "Nº", range(1, len(df_filtered) + 1))

    return df_filtered

def get_model_structure():
    """Lê o arquivo modelo MEDIÇÕES.xlsx para obter a ordem exata das colunas e estilos."""
    model_path = FILE_BASE.replace("BASE.xlsx", "MEDIÇÕES.xlsx")
    model_widths = {}
    model_header_style = {}
    ordered_columns = []
    
    try:
        wb_mod = openpyxl.load_workbook(model_path, data_only=False)
        ws_mod = wb_mod['Medições']
        
        # Ler cabeçalhos da linha 2
        for i in range(1, ws_mod.max_column + 1):
            cell_mod = ws_mod.cell(row=2, column=i)
            val_mod = cell_mod.value
            if val_mod:
                name_clean = str(val_mod).replace('\n', ' ').strip()
                ordered_columns.append(name_clean)
                
                col_let = get_column_letter(i)
                w = ws_mod.column_dimensions[col_let].width
                model_widths[name_clean] = w
                
                model_header_style[name_clean] = {
                    'fill': cell_mod.fill.start_color.rgb if cell_mod.fill else None,
                    'font_bold': cell_mod.font.bold if cell_mod.font else False,
                    'font_color': cell_mod.font.color.rgb if cell_mod.font and cell_mod.font.color else None
                }
    except Exception as e:
        print(f"Erro ao ler modelo: {e}")
        return [], {}, {}
    
    # Adicionar FISCAL se não houver no modelo (após GESTOR)
    if "GESTOR" in ordered_columns and "FISCAL" not in ordered_columns:
        idx = ordered_columns.index("GESTOR")
        ordered_columns.insert(idx + 1, "FISCAL")
        model_widths["FISCAL"] = 25
        # Copia estilo do GESTOR
        if "GESTOR" in model_header_style:
            model_header_style["FISCAL"] = model_header_style["GESTOR"].copy()
        
    return ordered_columns, model_widths, model_header_style

def main():
    print("Iniciando...")
    
    # 1. Obter estrutura do modelo
    ordered_columns, model_widths, model_header_style = get_model_structure()
    
    if not ordered_columns:
        print("ALERTA: Não foi possível ler colunas do modelo. Usando fallback.")
        return 

    # 2. Carregar mapeamentos
    region_map = get_region_mapping()
    comissoes_map = get_gestor_fiscal_data() # Agora unificado
    contractor_map = get_contractor_mapping()
    concluidas_sei: Any = get_concluidas_sei() # Novos SEIs para mover para PROBLEMAS

    # 3. Carregar DADOS
    df_ana = pd.read_excel(FILE_ANALITICA)
    df_ana['SEI_CLEAN'] = df_ana['Processo SEI'].apply(clean_sei)
    df_ana = df_ana.drop_duplicates(subset=['SEI_CLEAN']).copy()

    df_base = pd.read_excel(FILE_BASE)
    df_base['SEI_CLEAN'] = df_base['Processo SEI'].apply(clean_sei)
    # Suporte ao novo formato BASE.xlsx (coluna 'Valor') e ao formato antigo ('Valor das medições')
    if 'Valor' in df_base.columns:
        df_base['Valor'] = df_base['Valor'].apply(to_numeric)
    elif 'Valor das medições' in df_base.columns:
        df_base['Valor'] = df_base['Valor das medições'].apply(to_numeric)
    else:
        raise KeyError("Coluna de valor não encontrada no BASE.xlsx. Esperado: 'Valor' ou 'Valor das medições'.")

    # Mapeamento de meses: número inteiro -> abreviação
    meses_pt = {1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR", 5: "MAI", 6: "JUN",
                7: "JUL", 8: "AGO", 9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ"}
    # Mapeamento de nomes completos em português -> abreviação
    meses_nome_pt = {
        "JANEIRO": "JAN", "FEVEREIRO": "FEV", "MARÇO": "MAR", "MARCO": "MAR",
        "ABRIL": "ABR", "MAIO": "MAI", "JUNHO": "JUN", "JULHO": "JUL",
        "AGOSTO": "AGO", "SETEMBRO": "SET", "OUTUBRO": "OUT",
        "NOVEMBRO": "NOV", "DEZEMBRO": "DEZ"
    }
    def format_mes_ano(r):
        try:
            # Garante que 2025.0 ou "2025" vire 2025 e depois suffix "25"
            ano_val = int(float(str(r['Ano'])))
            ano_s = str(ano_val)
        except:
            ano_s = str(r['Ano']) # type: ignore
        suffix = ano_s[-2:] if len(ano_s) >= 2 else ano_s # type: ignore
        mes_raw = r['Mês']
        mes_str = "JAN"
        if pd.notna(mes_raw):
            mes_val_s = str(mes_raw).strip().upper()
            if mes_val_s in meses_nome_pt:
                # Nome completo: "Janeiro", "JANEIRO", etc.
                mes_str = meses_nome_pt[mes_val_s]
            else:
                # Fallback numérico: 1, 2, ..., 12
                try:
                    mes_str = meses_pt.get(int(float(mes_val_s)), "JAN")
                except (ValueError, TypeError):
                    mes_str = "JAN"
        return f"{mes_str}/{suffix}"
    df_base['MesAno'] = df_base.apply(format_mes_ano, axis=1)

    df_pivot = df_base.pivot_table(index='SEI_CLEAN', columns='MesAno', values='Valor', aggfunc='sum').fillna(0)

    # 4. Consolidar dados
    final_rows = []
    gestores_faltantes = []
    
    for _, row in df_ana.iterrows():
        sei = row['SEI_CLEAN']
        info = comissoes_map.get(sei, {'gestor': '', 'local': 'CIVIS', 'status_aux': ''})
        
        # Dados básicos
        fase_original = str(row['Fase']).strip().upper() if pd.notna(row.get('Fase')) and str(row['Fase']).strip() else info.get('status_aux', '')
        
        # Exceções: Força status "EXECUÇÃO" para o SEI solicitado
        if sei == "330018/000567/2021":
            status_final = "EXECUÇÃO"
        elif sei in concluidas_sei:
            status_final = "CONCLUÍDA"
        else:
            status_final = fase_original

        dados = {
            "SEI": row['Processo SEI'],
            "LOCAL": info['local'],
            "STATUS": status_final,
            "GESTOR": info['gestor'],
            "FISCAL": info.get('fiscal', ''),
            "MUNICIPIO": row['Municipio'],
            "REGIÃO": region_map.get(str(row['Municipio']).strip().upper(), ""),
            "CONTRATADA": contractor_map.get(normalize_name(str(row['Contratada'])), str(row['Contratada']).strip())
        }

        if not dados['GESTOR']:
            gestores_faltantes.append({'SEI': dados['SEI'], 'CONTRATADA': row['Contratada']})

        # Datas e Prazos
        dt_ini = pd.to_datetime(row['Ordem de Início'], errors='coerce')
        dt_fim = pd.to_datetime(row['Prazo Final'], errors='coerce')
        dados["ORDEM DE INÍCIO"] = dt_ini
        dados["DATA FINAL"] = dt_fim
        dados["PRAZO EXECUÇÃO"] = (dt_fim - dt_ini).days if pd.notnull(dt_fim) and pd.notnull(dt_ini) else ""

        # Financeiro
        vlr_contr = to_numeric(row['Valor contrato (Atual)'])
        dados["VLR.CONTRATO C/ADITIVO"] = vlr_contr
        
        # Novas fontes do ANALITICA.xlsx conforme pedido do usuário
        # "MEDIÇÕES ACUMULADAS" vem de "Valor contrato (Atual)"
        # "% EXEC" vem de "Acumulado atual (%)"
        # "SALDO DO CONTRATO" vem de "Saldo Atual do Contrato"
        vlr_acum_ana = to_numeric(row['Valor contrato (Atual)'])
        perc_exec_ana = to_numeric(row['Acumulado atual (%)'])
        saldo_ana = to_numeric(row['Saldo Atual do Contrato'])

        df_pivot_cols: Any = df_pivot.columns # type: ignore
        med_acum_calc: float = 0.0 # Valor calculado das mensais para conferência interna
        med_2025: float = 0.0
        med_2026: float = 0.0
        
        # Preencher meses baseado nas colunas do modelo
        for col_name in ordered_columns:
            # Tenta casar com padrão de mês
            col_clean = str(col_name).replace(" ", "")
            
            # Se a coluna existe no pivot, puxa o valor
            if col_clean in df_pivot_cols:
                val_raw = df_pivot.loc[sei, col_clean] if sei in df_pivot.index else 0.0 # type: ignore
                val: float = float(val_raw)
                dados[col_name] = float(round(val, 2)) # type: ignore
                med_acum_calc += val
                if "/25" in str(col_clean):
                    # Verifica se é coluna de mês para somar no MEDIÇÕES 2025
                    if re.match(r'^[A-Z]{3}/\d{2}$', str(col_clean)):
                        med_2025 += val # type: ignore
                elif "/26" in str(col_clean):
                    # Verifica se é coluna de mês para somar no MEDIÇÕES 2026
                    if re.match(r'^[A-Z]{3}/\d{2}$', str(col_clean)):
                        med_2026 += val # type: ignore
            elif col_name not in dados:
                 # Coluna do modelo que não é dado básico e não está no pivot
                 pass

        # Atribui conforme nova regra (ANALITICA.xlsx)
        dados["% EXEC."] = perc_exec_ana
        # Se % EXEC. for zero, não exibir o conteúdo de MEDIÇÕES ACUMULADAS
        if perc_exec_ana == 0 or perc_exec_ana == 0.0:
            dados["MEDIÇÕES ACUMULADAS"] = ""
        else:
            dados["MEDIÇÕES ACUMULADAS"] = vlr_acum_ana
        dados["MEDIÇÕES 2025"] = float(round(med_2025, 2)) # type: ignore
        dados["MEDIÇÕES 2026"] = float(round(med_2026, 2)) # type: ignore
        dados["SALDO DO CONTRATO"] = saldo_ana

        # Montar linha final ordenada
        linha_ordenada = {}
        for col in ordered_columns:
            val_out = dados.get(col)
            
            # Fallback de Nomes
            if val_out is None:
                if "PRAZO" in col and "EXECUÇÃO" in col: val_out = dados.get("PRAZO EXECUÇÃO")
                elif "ORDEM" in col and "INÍCIO" in col: val_out = dados.get("ORDEM DE INÍCIO")
                elif "VLR" in col and "CONTRATO" in col: val_out = dados.get("VLR.CONTRATO C/ADITIVO")
                elif "MEDIÇÕES" in col and "ACUMULADAS" in col: val_out = dados.get("MEDIÇÕES ACUMULADAS")
                elif "MEDIÇÕES" in col and "2025" in col: val_out = dados.get("MEDIÇÕES 2025")
                elif "MEDIÇÕES" in col and "2026" in col: val_out = dados.get("MEDIÇÕES 2026")
                elif "SALDO" in col and "CONTRATO" in col: val_out = dados.get("SALDO DO CONTRATO")
                elif "%" in col and "EXEC" in col: val_out = dados.get("% EXEC.")
            
            linha_ordenada[col] = val_out if val_out is not None else ""
            
        final_rows.append(linha_ordenada)

    df_all = pd.DataFrame(final_rows)
    # Garante a ordem das colunas
    df_all = df_all[ordered_columns]

    # Separar em EXECUÇÃO e PROBLEMAS
    df_execucao = prepare_dataframe(df_all, keep_execution=True)
    df_problemas = prepare_dataframe(df_all, keep_execution=False)

    # REMOVER FISCAL SOMENTE DA ABA MEDIÇÕES
    if "FISCAL" in df_execucao.columns:
        df_execucao = df_execucao.drop(columns=["FISCAL"])

    # Escrever
    with pd.ExcelWriter(FILE_OUTPUT, engine='openpyxl') as writer:
        df_execucao.to_excel(writer, sheet_name='Medições', index=False)
        if not df_problemas.empty:
            df_problemas.to_excel(writer, sheet_name='PROBLEMAS', index=False)
        if gestores_faltantes:
            pd.DataFrame(gestores_faltantes).to_excel(writer, sheet_name='GESTOR_FALTANTES', index=False)

    # Formatar
    wb = openpyxl.load_workbook(FILE_OUTPUT)
    
    # Define colunas por aba
    cols_medicoes = [c for c in ordered_columns if c != "FISCAL"]
    cols_problemas = list(ordered_columns)

    sheet_configs = {
        'Medições': cols_medicoes,
        'PROBLEMAS': cols_problemas
    }

    for sheet_name, sheet_cols in sheet_configs.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            apply_sheet_formatting(ws, 
                                   col_map={c: i+1 for i, c in enumerate(sheet_cols)}, 
                                   header=sheet_cols, 
                                   all_months=[], 
                                   model_widths=model_widths, 
                                   model_header_style=model_header_style,
                                   h_vlr_contr="VLR.CONTRATO C/ADITIVO", 
                                   h_med_acum="MEDIÇÕES ACUMULADAS", 
                                   h_saldo="SALDO DO CONTRATO", 
                                   h_inicio="ORDEM DE INÍCIO")

    wb.save(FILE_OUTPUT)
    print(f"Finalizado: {FILE_OUTPUT}")
    print(f"  - Aba 'Medições': {len(df_execucao)} obras em EXECUÇÃO")
    print(f"  - Aba 'PROBLEMAS': {len(df_problemas)} obras com status != EXECUÇÃO")
    if gestores_faltantes:
        print(f"  - Aba 'GESTOR_FALTANTES': {len(gestores_faltantes)} registros sem gestor")

if __name__ == "__main__":
    main()
