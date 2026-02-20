import openpyxl

FILE_MODELO = r"d:\APRENDIZADO APP\MEDICOES\MEDIÇÕES_CONSOLIDADO.xlsx"
wb = openpyxl.load_workbook(FILE_MODELO)
ws = wb['Medições']

widths = {}
for i in range(1, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(i)
    width = ws.column_dimensions[col_letter].width
    header_val = ws.cell(row=1, column=i).value # Headers are in row 1 in Consolidated file
    if header_val:
        widths[str(header_val).replace('\n', ' ')] = width

for k, v in widths.items():
    print(f"'{k}': {v}")
