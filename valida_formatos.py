import openpyxl

FILE_OUTPUT = r"d:\APRENDIZADO APP\MEDIÇÕES\MEDIÇÕES_CONSOLIDADO.xlsx"
wb = openpyxl.load_workbook(FILE_OUTPUT)
ws = wb['Medições']

# Check cell G2 (VLR.CONTRATO C/ADITIVO)
cell = ws['G2']
print(f"Cell G2 value: {cell.value}")
print(f"Cell G2 number_format: {cell.number_format}")

# Check cell P2 (% EXEC.)
cell = ws['P2']
print(f"Cell P2 value: {cell.value}")
print(f"Cell % EXEC. number_format: {cell.number_format}")

# Check cell R2 (JAN/21)
cell = ws['R2']
print(f"Cell R2 value: {cell.value}")
print(f"Cell R2 number_format: {cell.number_format}")
