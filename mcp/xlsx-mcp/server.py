from __future__ import annotations

from typing import Any, Dict, List, Optional, Union
import os
import pandas as pd

from mcp.server.fastmcp import FastMCP

# MCP Server name (aparece no host)
mcp = FastMCP("xlsx-mcp")

def _resolve_path(path: str) -> str:
    """Resolve caminho absoluto e valida existência."""
    p = os.path.abspath(os.path.expanduser(path))
    if not os.path.exists(p):
        raise FileNotFoundError(f"Arquivo não encontrado: {p}")
    if not p.lower().endswith(".xlsx"):
        raise ValueError("O arquivo deve ser .xlsx")
    return p

def _read_sheet(path: str, sheet: Union[str, int], header: Optional[int] = 0) -> pd.DataFrame:
    """Lê uma aba do XLSX em DataFrame."""
    p = _resolve_path(path)
    return pd.read_excel(p, sheet_name=sheet, header=header, engine="openpyxl")

def _to_records(df: pd.DataFrame, max_rows: int) -> List[Dict[str, Any]]:
    """Converte DataFrame para lista de registros (JSON-friendly)."""
    if max_rows is not None and max_rows > 0:
        df = df.head(max_rows)
    # Substitui NaN por None para JSON limpo
    return df.where(pd.notnull(df), None).to_dict(orient="records")

@mcp.tool()
def xlsx_list_sheets(path: str) -> Dict[str, Any]:
    """
    Lista as abas disponíveis em um arquivo XLSX.
    Args:
      path: caminho do arquivo .xlsx (local).
    """
    p = _resolve_path(path)
    xls = pd.ExcelFile(p, engine="openpyxl")
    return {"path": p, "sheets": xls.sheet_names}

@mcp.tool()
def xlsx_preview(path: str, sheet: Union[str, int] = 0, max_rows: int = 25) -> Dict[str, Any]:
    """
    Mostra um preview (primeiras linhas) de uma aba do XLSX.
    Args:
      path: caminho do .xlsx
      sheet: nome da aba ou índice (0 = primeira)
      max_rows: quantidade máxima de linhas no preview
    """
    df = _read_sheet(path, sheet=sheet, header=0)
    return {
        "path": os.path.abspath(path),
        "sheet": sheet,
        "rows": _to_records(df, max_rows=max_rows),
        "columns": list(df.columns),
        "row_count": int(df.shape[0]),
        "col_count": int(df.shape[1]),
    }

@mcp.tool()
def xlsx_get_cell(path: str, sheet: Union[str, int], row_1based: int, col_1based: int) -> Dict[str, Any]:
    """
    Lê uma célula por posição (1-based), como no Excel.
    Args:
      row_1based: linha (1 = primeira linha da planilha)
      col_1based: coluna (1 = coluna A)
    Observação:
      Este método lê via openpyxl (não via pandas) para respeitar posição de célula.
    """
    from openpyxl import load_workbook

    p = _resolve_path(path)
    wb = load_workbook(p, data_only=True, read_only=True)
    ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]
    value = ws.cell(row=row_1based, column=col_1based).value
    return {
        "path": p,
        "sheet": sheet,
        "row": row_1based,
        "col": col_1based,
        "value": value
    }

@mcp.tool()
def xlsx_summarize(path: str, sheet: Union[str, int] = 0, max_rows_scan: int = 200) -> Dict[str, Any]:
    """
    'Interpreta' a aba: tipos de colunas, nulos, estatísticas básicas e amostra.
    Útil para o agente entender a planilha automaticamente.
    """
    df = _read_sheet(path, sheet=sheet, header=0)
    scan = df.head(max_rows_scan)

    summary = []
    for c in scan.columns:
        s = scan[c]
        summary.append({
            "column": str(c),
            "dtype": str(s.dtype),
            "nulls": int(s.isna().sum()),
            "example_values": [v for v in s.dropna().head(5).tolist()]
        })

    numeric_cols = [c for c in scan.columns if pd.api.types.is_numeric_dtype(scan[c])]
    stats = {}
    if numeric_cols:
        stats = scan[numeric_cols].describe(include="all").where(pd.notnull, None).to_dict()

    return {
        "path": os.path.abspath(path),
        "sheet": sheet,
        "columns_summary": summary,
        "numeric_stats": stats,
        "sample_rows": _to_records(scan, max_rows=10),
        "row_count": int(df.shape[0]),
        "col_count": int(df.shape[1]),
    }

if __name__ == "__main__":
    # FastMCP roda via stdio quando chamado pelo host,
    # mas também permite executar diretamente em modo de desenvolvimento.
    mcp.run()
